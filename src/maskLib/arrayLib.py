# -*- coding: utf-8 -*-
"""
Created 2026

@author: Agrim

arrayLib: parameter-sweep ("dose array") machinery for field-grid layouts.

CONCEPTS
========
A chip is treated as a grid of identical square FIELDS (e.g. 40 x 40 fields
of 500 um), each holding one device. A sweep varies device parameters from
field to field so a single chip tests many parameter combinations.

Two sweep classes are provided:

  Sweep2D -- two parameters axes over the whole grid:
                 col: varies along x (grid_nx steps)
                 row: varies along y (grid_ny steps)

  Sweep3D -- the grid is tiled with identical tile_nx x tile_ny blocks
             ("tiles"), adding a third axis:
                 col:  varies along x inside a tile (tile_nx steps)
                 row:  varies along y inside a tile (tile_ny steps)
                 tile: varies from tile to tile     (one step per tile)
             Example: a 10 x 20 tile in a 40 x 40 grid gives 4 x 2 = 8
             tiles, so the tile axis has 8 steps. Tile size must divide the
             grid evenly (a Sweep2D is just a Sweep3D whose tile IS the grid).

AXIS SPECIFICATION
==================
Each axis is a dict {parameter_name: (start, step)}:

    SWEEP_COL = {'smallfinger_width': (0.100, 0.010)}   # 0.10, 0.11, 0.12, ...

The number of steps is fixed by the axis (tile_nx, tile_ny, or the number
of tiles) and the FINAL VALUE IS COMPUTED AUTOMATICALLY:

    final = start + (steps - 1) * step

Values are rounded to 9 decimals so clean steps stay decimal-exact (no
0.30000000000000004). Several parameters may share one axis -- they sweep
in lockstep (e.g. keep bigfinger_width = smallfinger_width + 0.2 by giving
both the same step).

PARAMETER KINDS
===============
What a parameter *does* is declared by two registries passed at creation:

    geometry_params = {'smallfinger_width': 'smallfingerW', ...}
        geometry ("Size") parameters change the drawn shapes; the value
        maps to the drawing kwarg named by the registry entry.

    dose_params = {'bridge_dose': 'BRIDGE', ...}
        dose parameters redirect a shape onto a per-dose layer named
        FAMILY_<value> (e.g. BRIDGE_400), so the ebeam writer can assign
        each dose its own exposure. All needed layers are enumerated by
        .dose_layers() -- add them to the wafer automatically, never by
        hand. Use the module-level dose_layer() helper when drawing.

LABELS
======
Every field gets a compact label:  <tile letter><column 2 digits><row 2 digits>
e.g. 'A0103' = tile A, column 01, row 03. A Sweep2D has a single tile, so
all its labels start with 'A'. Columns/rows are zero-padded to two digits
('00'-'09', '10', ...).

TYPICAL USAGE
=============
    from maskLib.arrayLib import Sweep3D, dose_layer

    sweep = Sweep3D(40, 40, 10, 20,
                    col={'smallfinger_width': (0.100, 0.010)},
                    row={'bridge_dose': (400, 40)},
                    tile={'smallfinger_dose': (800, 100)},
                    geometry_params=GEOMETRY_PARAMS, dose_params=DOSE_PARAMS)
    sweep.print_summary()                       # human-readable sweep report
    layers = sweep.dose_layers()                # auto per-dose layer names
    params, label = sweep.field(ix, iy)         # what to draw at field (ix,iy)
    lyr = dose_layer('BRIDGE', params, 'bridge_dose')   # layer for a shape
    sweep.export_workbook('mysweep.xlsx')       # parameter table + maps
"""

import numpy as np


def fmt_value(v):
    '''compact number formatting for layer names and labels (400.0 -> '400')'''
    return '%g' % v


def index_letter(i):
    '''0->'A' ... 25->'Z', 26->'AA', ... (spreadsheet column style)'''
    s = ''
    while True:
        s = chr(ord('A') + i % 26) + s
        i = i // 26 - 1
        if i < 0:
            return s


def dose_layer(base, params, dose_name):
    '''
    Layer name for one shape of one field: BASE if that shape's dose is not
    being swept (dose_name absent from params), else BASE_<dose>.

        dose_layer('BRIDGE', {'bridge_dose': 400}, 'bridge_dose') -> 'BRIDGE_400'
        dose_layer('BRIDGE', {}, 'bridge_dose')                   -> 'BRIDGE'
    '''
    v = params.get(dose_name)
    return base if v is None else '%s_%s' % (base, fmt_value(v))


def export_ldt(path, entries):
    '''
    Write an Elionix layer dose table (.ldt) for ebeam lithography.

    entries: iterable of (layer_number, dose) pairs. layer_number is the
    numeric layer the DXF layer becomes after GDS conversion (= its index
    in the wafer layer table, wafer.layerNums[name]). The dose is written
    divided by 1000 with 3 decimals, per the Elionix format:

        Dosetable V1.0
        Dose Assignment by Layer
        (2 ,   0.400)
        ...

    Returns the path written.
    '''
    with open(path, 'w') as f:
        f.write('\nDosetable V1.0\nDose Assignment by Layer\n')
        for num, dose in sorted(entries):
            f.write('(%d ,   %.3f)\n' % (num, dose / 1000.0))
        f.write('\n')
    return path


class Sweep3D:
    '''
    3D parameter sweep over a tiled field grid (see module docstring).

    grid_nx, grid_ny -- the reference field grid (the main chiplet)
    tile_nx, tile_ny -- tile size in fields; must divide the grid evenly
    col, row, tile   -- axis dicts {parameter_name: (start, step)}
    geometry_params  -- registry {parameter_name: drawing_kwarg}
    dose_params      -- registry {parameter_name: LAYER_FAMILY}

    Secondary chips with a different grid (e.g. smaller corner chips) can
    reuse the same sweep through field(..., grid_nx=..., grid_ny=...,
    strict=False); their tile parameter values wrap around.
    '''

    def __init__(self, grid_nx, grid_ny, tile_nx, tile_ny,
                 col=None, row=None, tile=None,
                 geometry_params=None, dose_params=None):
        assert grid_nx % tile_nx == 0 and grid_ny % tile_ny == 0, (
            'Tile size %dx%d does not evenly fill the %dx%d field grid -- '
            'pick tile dimensions that divide the grid'
            % (tile_nx, tile_ny, grid_nx, grid_ny))
        self.grid_nx, self.grid_ny = grid_nx, grid_ny
        self.tile_nx, self.tile_ny = tile_nx, tile_ny
        self.tiles_x = grid_nx // tile_nx    # tiles across
        self.tiles_y = grid_ny // tile_ny    # tiles up
        self.n_tiles = self.tiles_x * self.tiles_y
        self.geometry_params = geometry_params or {}
        self.dose_params = dose_params or {}

        # expand each axis {param: (start, step)} into {param: value array};
        # rounding keeps clean steps decimal-exact (no float noise)
        self.axes = {}
        for axname, spec, steps in (('col', col, tile_nx),
                                    ('row', row, tile_ny),
                                    ('tile', tile, self.n_tiles)):
            values = {}
            for pname, (start, step) in (spec or {}).items():
                assert pname in self.geometry_params or pname in self.dose_params, (
                    'unknown sweep parameter: %s (add it to geometry_params '
                    'or dose_params)' % pname)
                values[pname] = np.round(start + step * np.arange(steps), 9)
            self.axes[axname] = values

    # ---- per-field lookups ---------------------------------------------

    def field(self, ix, iy, grid_nx=None, grid_ny=None, strict=True):
        '''
        (params, label) for field (ix, iy).

        params -- {parameter_name: value} with one entry per swept parameter
        label  -- e.g. 'A0103' = tile A, column 01, row 03

        grid_nx/grid_ny default to the reference grid; pass them (with
        strict=False) for secondary chips such as corner chips. strict=True
        asserts that tiles fill the grid evenly; with strict=False the tile
        parameter values simply wrap around (tile index modulo n_tiles).
        '''
        gx = self.grid_nx if grid_nx is None else grid_nx
        gy = self.grid_ny if grid_ny is None else grid_ny
        if strict:
            assert gx % self.tile_nx == 0 and gy % self.tile_ny == 0, (
                'Tile size %dx%d does not evenly fill the %dx%d field grid'
                % (self.tile_nx, self.tile_ny, gx, gy))
        ti = ix % self.tile_nx          # column inside the tile
        tj = iy % self.tile_ny          # row inside the tile
        tile = (ix // self.tile_nx) + (iy // self.tile_ny) * (gx // self.tile_nx)
        params = {}
        for pname, vals in self.axes['col'].items():
            params[pname] = vals[ti]
        for pname, vals in self.axes['row'].items():
            params[pname] = vals[tj]
        for pname, vals in self.axes['tile'].items():
            params[pname] = vals[tile % len(vals)]
        return params, '%s%02d%02d' % (index_letter(tile), ti, tj)

    def param_names(self):
        '''sorted names of every swept parameter (all axes combined)'''
        return sorted({p for ax in self.axes.values() for p in ax})

    def kind(self, pname):
        ''''geometry' or 'dose' for one parameter'''
        return 'geometry' if pname in self.geometry_params else 'dose'

    def dose_layers(self):
        '''
        Every per-dose layer name implied by the sweep, in a stable order,
        e.g. ['BRIDGE_400', 'BRIDGE_440', ..., 'SMALLFINGER_800', ...].
        Feed these to wafer.SetupLayers -- never write dose layers by hand.
        '''
        out = []
        for axname in ('col', 'row', 'tile'):
            for pname, vals in self.axes[axname].items():
                if pname in self.dose_params:
                    for v in vals:
                        lyr = '%s_%s' % (self.dose_params[pname], fmt_value(v))
                        if lyr not in out:
                            out.append(lyr)
        return out

    def ldt_entries(self, layer_number, base_doses=None):
        '''
        (layer_number, dose) pairs for an Elionix dose table (see export_ldt).

        layer_number -- function mapping a layer NAME to its numeric layer,
                        e.g.  lambda name: wafer.layerNums[name]
        base_doses   -- {LAYER_FAMILY: dose} for ebeam layers whose dose is
                        NOT being swept (e.g. {'LEADS': 1200}). Families that
                        ARE swept are ignored here -- they get one entry per
                        auto-generated dose layer instead, with the dose
                        taken straight from the layer's own sweep value.
        '''
        entries = []
        swept_families = set()
        for axname in ('col', 'row', 'tile'):
            for pname, vals in self.axes[axname].items():
                if pname in self.dose_params:
                    fam = self.dose_params[pname]
                    swept_families.add(fam)
                    for v in vals:
                        entries.append((layer_number('%s_%s' % (fam, fmt_value(v))),
                                        float(v)))
        for fam, dose in (base_doses or {}).items():
            if fam not in swept_families:
                entries.append((layer_number(fam), float(dose)))
        return entries

    # ---- reporting -------------------------------------------------------

    def axis_kind(self, axname):
        '''
        "Size" if the axis sweeps geometry, "Dose" if it sweeps doses,
        "Mixed" if both, "None" if the axis is unused.
        '''
        kinds = {('Size' if p in self.geometry_params else 'Dose')
                 for p in self.axes[axname]}
        if not kinds:
            return 'None'
        return kinds.pop() if len(kinds) == 1 else 'Mixed'

    def sweep_type(self):
        '''
        Compact sweep classification, one word per active axis in
        col-row-tile order: e.g. 'SizeDoseDose' (3D) or 'SizeDose' (2D).
        '''
        return ''.join(k for k in (self.axis_kind(ax) for ax in ('col', 'row', 'tile'))
                       if k != 'None')

    def print_summary(self):
        '''print a full human-readable description of the sweep'''
        dims = '3D' if self.axes['tile'] else '2D'
        print('=' * 68)
        print('%s parameter sweep: %s  (col / row%s)'
              % (dims, self.sweep_type(), ' / tile' if self.axes['tile'] else ''))
        print('Field grid %d x %d; tile %d x %d -> %d x %d = %d tile(s) (%s-%s)'
              % (self.grid_nx, self.grid_ny, self.tile_nx, self.tile_ny,
                 self.tiles_x, self.tiles_y, self.n_tiles,
                 index_letter(0), index_letter(self.n_tiles - 1)))
        for axname, along, steps in (('col', 'along x in tile', self.tile_nx),
                                     ('row', 'along y in tile', self.tile_ny),
                                     ('tile', 'per tile', self.n_tiles)):
            if not self.axes[axname]:
                continue
            print('%s axis (%d steps, %s):' % (axname.upper(), steps, along))
            for pname, vals in self.axes[axname].items():
                step = vals[1] - vals[0] if len(vals) > 1 else 0
                if self.kind(pname) == 'dose':
                    target = 'dose -> %s_* layers' % self.dose_params[pname]
                else:
                    target = 'geometry -> %s' % self.geometry_params[pname]
                print('    %-20s %8s -> %-8s step %-6s (%s)'
                      % (pname, fmt_value(vals[0]), fmt_value(vals[-1]),
                         fmt_value(step), target))
        n = len(self.dose_layers())
        if n:
            print('Auto-generated dose layers: %d' % n)
        print('=' * 68)

    def legend_lines(self):
        '''compact per-parameter summary lines for drawing in a chip margin'''
        lines = []
        for tag, axname in (('COLS', 'col'), ('ROWS', 'row'), ('TILES', 'tile')):
            for pname, vals in self.axes[axname].items():
                step = vals[1] - vals[0] if len(vals) > 1 else 0
                lines.append('%s: %s = %s to %s step %s'
                             % (tag, pname, fmt_value(vals[0]),
                                fmt_value(vals[-1]), fmt_value(step)))
        return lines

    # ---- export ----------------------------------------------------------

    def export_workbook(self, path, grid_nx=None, grid_ny=None, strict=True):
        '''
        Write an .xlsx workbook (requires openpyxl) with:

          'parameters'  -- one row per field: label, ix, iy, and the value
                           of every swept parameter (for analysis scripts)
          'map'         -- field labels laid out in cells matching their
                           position on the chip (top row = top of chip)
          one sheet per swept parameter -- its values laid out like the
                           chip, with a white->red min-to-max color gradient
                           so the sweep pattern is visible at a glance

        Covers the reference grid by default; pass grid_nx/grid_ny (and
        strict=False) to export a secondary chip's smaller grid, e.g. a
        20 x 20 corner chip. Sheet names are the parameter names (truncated
        to Excel's 31-char limit). Returns the path written.
        '''
        from openpyxl import Workbook
        from openpyxl.formatting.rule import ColorScaleRule
        from openpyxl.utils import get_column_letter

        gx = self.grid_nx if grid_nx is None else grid_nx
        gy = self.grid_ny if grid_ny is None else grid_ny

        def lookup(ix, iy):
            return self.field(ix, iy, grid_nx=gx, grid_ny=gy, strict=strict)

        pnames = self.param_names()
        wb = Workbook()
        ws = wb.active
        ws.title = 'parameters'
        ws.append(['label', 'ix', 'iy'] + pnames)
        for ix in range(gx):
            for iy in range(gy):
                params, flabel = lookup(ix, iy)
                ws.append([flabel, ix, iy] + [float(params[p]) for p in pnames])

        def grid_sheet(title, cellvalue):
            '''new sheet laid out like the chip: columns = ix, rows = iy
            top-down (so the top spreadsheet row is the top of the chip)'''
            s = wb.create_sheet(title)
            s.append(['iy\\ix'] + list(range(gx)))
            for iy in reversed(range(gy)):
                s.append([iy] + [cellvalue(ix, iy) for ix in range(gx)])
            return s

        grid_sheet('map', lambda ix, iy: lookup(ix, iy)[1])

        # one gradient-colored value map per swept parameter
        data_range = 'B2:%s%d' % (get_column_letter(gx + 1), gy + 1)
        for pname in pnames:
            s = grid_sheet(pname[:31],
                           lambda ix, iy, p=pname: float(lookup(ix, iy)[0][p]))
            s.conditional_formatting.add(data_range, ColorScaleRule(
                start_type='min', start_color='FFFFFFFF',
                end_type='max', end_color='FFFF4444'))
        wb.save(path)
        return path


class Sweep2D(Sweep3D):
    '''
    2D parameter sweep: parameters vary along the columns and rows of the
    WHOLE grid -- no tiling. Implemented as a Sweep3D whose single tile is
    the entire grid, so every method (field, dose_layers, print_summary,
    export_workbook, ...) works identically. All labels start with tile
    letter 'A'.

        sweep = Sweep2D(40, 40,
                        col={'smallfinger_width': (0.100, 0.005)},  # 40 steps
                        row={'bridge_dose': (400, 20)},             # 40 steps
                        geometry_params=..., dose_params=...)
    '''

    def __init__(self, grid_nx, grid_ny, col=None, row=None,
                 geometry_params=None, dose_params=None):
        Sweep3D.__init__(self, grid_nx, grid_ny, grid_nx, grid_ny,
                         col=col, row=row, tile=None,
                         geometry_params=geometry_params,
                         dose_params=dose_params)
